# -*- coding: utf-8 -*-
"""gen_knights.py — LSO Knights LIST A: Analysis (Excel) + Runbook (PDF), new-standard pipeline.

Analysis = .xlsx (Matchups / Bands / Tapestry / Profiles). Runbook = PDF. Pulls the tapestry/units/
per-matchup narrative from lso_data and the LIVE win% + forced-mission + opponent disposition from the
real-mission sim (mc_knights_sim, 10k games). Verdicts DERIVED from the simulated win%.

  PYTHONPATH=tools:src python3 tools/gen_knights.py
"""
import os
import lso_data as L
import mc_knights_sim as S
import gen_pdf as G
import doc_versions as V
import sim_game
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

GAMES, SEED = 10000, 11
HEAD = PatternFill("solid", fgColor="1f3a5f")           # Knights steel-blue
HFONT = Font(bold=True, color="FFFFFF")
WRAP = Alignment(wrap_text=True, vertical="top")
VF = {"fav": "C6EFCE", "even": "FFEB9C", "unfav": "FFC7CE"}


def _verdict(w):
    if w >= 62:
        return "Favourable", "fav"
    if w >= 55:
        return "Lean favourable", "fav"
    if w >= 45:
        return "Coin-flip", "even"
    if w >= 40:
        return "Lean unfavourable", "unfav"
    return "Unfavourable", "unfav"


def _rows():
    sim = {r["mkey"]: r for r in S.results(GAMES, SEED)}
    tot = wsum = 0
    rows = []
    for m in L.MATCHUPS:
        r = sim.get(m["key"])
        if not r:
            continue
        w = r["win"]; tot += r["prev"]; wsum += r["prev"] * w
        verdict, cls = _verdict(w)
        rows.append(dict(m=m, win=w, prev=r["prev"], disp=r["disp"], mission=r["mission"],
                         opp_mission=r["opp_mission"], verdict=verdict, cls=cls))
    return rows, round(wsum / tot)


def _dn(k):
    return sim_game.DISP_NAME.get(k, k)


def _banner(ws, text, ncol, size=11):
    r = ws.max_row + (1 if ws.max_row > 1 else 0) + 1
    c = ws.cell(r, 1, text); c.font = Font(bold=True, size=size, color="FFFFFF"); c.fill = HEAD; c.alignment = WRAP
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=ncol)
    return r


def _row(ws, cells, fills=None, bold=False):
    r = ws.max_row + 1
    for i, val in enumerate(cells, 1):
        c = ws.cell(r, i, val); c.alignment = WRAP
        if bold:
            c.font = Font(bold=True)
        if fills and i - 1 < len(fills) and fills[i - 1]:
            c.fill = PatternFill("solid", fgColor=fills[i - 1])


def analysis_xlsx(rows, weighted):
    wb = Workbook(); ws = wb.active; ws.title = "Matchups"
    ws.column_dimensions["A"].width = 46
    for col, w in [("B", 16), ("C", 18), ("D", 8), ("E", 16)]:
        ws.column_dimensions[col].width = w
    _banner(ws, L.LIST_A_NAME, 5, 12)
    for line in [L.DETACHMENTS, L.DISPOSITION,
                 f"Prevalence-weighted win rate: ~{weighted}%.  FINDING: disposition Purge the Foe splits "
                 "the field — Knights FEAST on Priority-Assets armies (they get Destroyer's Wrath, a kill "
                 "mission) but STARVE vs Take-and-Hold armies (Unstoppable Force is a control mission a "
                 "~6-model army can't play). The real bogeys are anti-tank Purge lists (Necrons gauss, T'au "
                 "rail). The mission-grounded read is harsher than the old damage-only verdicts — Knights "
                 "out-KILL but can't out-SCORE the durable objective-holders."]:
        _row(ws, [line], bold=True)
        ws.merge_cells(start_row=ws.max_row, start_column=1, end_row=ws.max_row, end_column=5)
    _banner(ws, "Matchups (Knights = Purge the Foe)", 5)
    _row(ws, ["Faction / archetype", "Opp disposition", "You play", "Win%", "Read"], bold=True)
    for cell in ws[ws.max_row]:
        cell.fill = HEAD; cell.font = HFONT
    for r in rows:
        m = r["m"]; fill = VF.get(r["cls"])
        _row(ws, [f'{m["faction"]} — {m["archetype"]}', _dn(r["disp"]), r["mission"], f'{r["win"]}%', r["verdict"]],
             fills=[None, None, None, fill, fill])

    wb2 = wb.create_sheet("Bands")
    wb2.column_dimensions["A"].width = 22; wb2.column_dimensions["B"].width = 90
    _banner(wb2, "Bands (from simulated win%)", 2)
    for label, lo, hi, key in [("Favourable (55%+)", 55, 101, "fav"), ("Coin-flip (45-54%)", 45, 55, "even"),
                               ("Unfavourable (<45%)", 0, 45, "unfav")]:
        g = sorted([x for x in rows if lo <= x["win"] < hi], key=lambda x: -x["win"])
        _row(wb2, [label, ", ".join(f'{x["m"]["faction"]} {x["win"]}%' for x in g) or "—"],
             fills=[VF[key], None], bold=True)

    wt = wb.create_sheet("Tapestry")
    wt.column_dimensions["A"].width = 30; wt.column_dimensions["B"].width = 100
    _banner(wt, "The rules tapestry (DB-verified)", 2)
    _row(wt, ["Rule", "What it does"], bold=True)
    for n, t in L.RULES:
        _row(wt, [n, t])

    wp = wb.create_sheet("Profiles")
    for col, w in [("A", 40), ("B", 42), ("C", 52)]:
        wp.column_dimensions[col].width = w
    _banner(wp, "Verified profiles", 3)
    for row in L.VERIFIED_PROFILES:
        _row(wp, [str(x) for x in row])
    _banner(wp, "Bottom line", 3)
    _row(wp, [L.RECORD_NOTE]); wp.merge_cells(start_row=wp.max_row, start_column=1, end_row=wp.max_row, end_column=3)

    V.stamp_xlsx(wb, "knights-analysis")
    out = V.out_path("knights-analysis"); os.makedirs(os.path.dirname(out), exist_ok=True)
    wb.save(out)
    return out


def runbook(rows, weighted):
    S1 = G.section("Mindset", G.p(G.esc(L.MINDSET)) + G.finding(
        f"Weighted ~{weighted}%. The disposition (Purge the Foe) decides the shape of your day: you WANT "
        "Priority-Assets opponents (you get Destroyer's Wrath — kill for VP) and you DREAD Take-and-Hold "
        "opponents (Unstoppable Force asks you to hold objectives a 6-model army can't). Kill their "
        "scoring units to steal the control you can't muscle."))
    tap = G.table(["Rule", "What it does"], [[G.esc(n), G.esc(t)] for n, t in L.RULES])
    S2 = G.section("Tapestry quick-reference", tap)

    plans = ""
    for r in rows:
        m = r["m"]
        head = (f'{m["faction"]} — <span class="{r["cls"]}">{r["win"]}% ({G.esc(r["verdict"])})</span> · '
                f'you play <b>{G.esc(r["mission"])}</b> vs their {_dn(r["disp"])} (they play {G.esc(r["opp_mission"])})')
        body = G.p(f'<b>Deciding factor:</b> {G.esc(m.get("deciding",""))}')
        for label, key in [("Mission / heist", "heist"), ("Kill priority", "kill_priority"), ("Deploy", "deploy")]:
            if m.get(key):
                body += G.p(f'<b>{label}:</b> {G.esc(m[key])}')
        plans += G.sub(head, body)
    S3 = G.section("Per-matchup battle plans (28 archetypes)", plans)

    units = G.table(["Unit", "×", "Wargear", "Pts", "Note"],
                    [[G.esc(u[0]), G.esc(u[1]), G.esc(u[2]), G.esc(u[3]), G.esc(u[4] if len(u) > 4 else "")]
                     for u in L.LIST_UNITS])
    S4 = G.section("The list (List A)", units)
    return [S1, S2, S3, S4]


def main():
    rows, weighted = _rows()
    print("wrote", analysis_xlsx(rows, weighted))
    print("wrote", G.render("knights-runbook", runbook(rows, weighted)))


if __name__ == "__main__":
    main()
