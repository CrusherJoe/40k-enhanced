# -*- coding: utf-8 -*-
"""gen_custodes.py — Analysis + Runbook PDFs for the Custodes "Better Thing 2" showcase.

PDF-only (skips Word). Pulls the tapestry/units/plans from custodes_data.py and the LIVE win% +
forced-mission + opponent disposition from the rebuilt real-mission sim (mc_custodes_sim). Verdicts
are DERIVED from the simulated win% so the docs never drift from the sim.

  PYTHONPATH=tools:src python3 tools/gen_custodes.py
"""
import os
import custodes_data as D
import mc_custodes_sim as S
import gen_pdf as G
import doc_versions as V
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

GAMES, SEED = 6000, 11
HEAD = PatternFill("solid", fgColor="4A3D10")
HFONT = Font(bold=True, color="FFFFFF")
WRAP = Alignment(wrap_text=True, vertical="top")
VF = {"fav": "C6EFCE", "even": "FFEB9C", "unfav": "FFC7CE"}


def _rows():
    """Join sim results with the narrative matchups (same faction order)."""
    sim = S.results(GAMES, SEED)
    tot = wsum = 0
    rows = []
    for r, m in zip(sim, D.MATCHUPS):
        w = r["win"]
        tot += r["prev"]; wsum += r["prev"] * w
        verdict, cls = _verdict(w)
        rows.append(dict(m=m, win=w, disp=r["disp"], mission=r["cu_mission"],
                         opp_mission=r["op_mission"], verdict=verdict, cls=cls))
    return rows, round(wsum / tot)


def _verdict(w):
    if w >= 62:
        return "Favourable", "fav"
    if w >= 55:
        return "Lean favourable", "fav"
    if w >= 45:
        return "Coin-flip", "even"
    if w >= 40:
        return "Lean unfavourable", "unfav"
    return "Unfavourable", "unfav"


def _disp_name(key):
    return {"take-and-hold": "Take and Hold", "purge-the-foe": "Purge the Foe",
            "reconnaissance": "Reconnaissance", "priority-assets": "Priority Assets",
            "disruption": "Disruption"}.get(key, key)


def _banner(ws, text, ncol, size=11):
    r = ws.max_row + (1 if ws.max_row > 1 else 0) + 1
    c = ws.cell(r, 1, text); c.font = Font(bold=True, size=size, color="FFFFFF"); c.fill = HEAD; c.alignment = WRAP
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=ncol)
    return r


def _row(ws, cells, fills=None, bold=False):
    r = ws.max_row + 1
    for i, val in enumerate(cells, 1):
        c = ws.cell(r, i, val); c.alignment = WRAP
        if bold:
            c.font = Font(bold=True)
        if fills and i - 1 < len(fills) and fills[i - 1]:
            c.fill = PatternFill("solid", fgColor=fills[i - 1])
    return r


def analysis_xlsx(rows, weighted):
    """The Analysis is an Excel workbook (Matchups / Bands / Tapestry / Profiles)."""
    wb = Workbook()
    ws = wb.active; ws.title = "Matchups"
    ws.column_dimensions["A"].width = 30
    for col in "BCDE":
        ws.column_dimensions[col].width = 20
    ws.column_dimensions["E"].width = 16

    _banner(ws, D.LIST_NAME, 5, 12)
    _row(ws, [D.DETACHMENTS]); ws.merge_cells(start_row=ws.max_row, start_column=1, end_row=ws.max_row, end_column=5)
    _row(ws, [D.DISPOSITION]); ws.merge_cells(start_row=ws.max_row, start_column=1, end_row=ws.max_row, end_column=5)
    _row(ws, [f"Prevalence-weighted win rate: ~{weighted}%.  FINDING: the biggest drag is the Purge-the-Foe "
              "matchups (Emperor's Children, T'au) — the matrix forces Priority-Assets Custodes onto Vital "
              "Link (hard) while they get Destroyer's Wrath. You out-fight them but lose the mission."], bold=True)
    ws.merge_cells(start_row=ws.max_row, start_column=1, end_row=ws.max_row, end_column=5)

    _banner(ws, "Matchups (Custodes = Priority Assets)", 5)
    _row(ws, ["Faction / archetype", "Opp disposition", "You play", "Win%", "Read"], bold=True)
    for cell in ws[ws.max_row]:
        cell.fill = HEAD; cell.font = HFONT
    for r in rows:
        m = r["m"]
        fill = VF.get(r["cls"])
        _row(ws, [f'{m["faction"]} — {m["archetype"]}', _disp_name(r["disp"]), r["mission"],
                  f'{r["win"]}%', r["verdict"]], fills=[None, None, None, fill, fill])

    # Bands sheet
    wb2 = wb.create_sheet("Bands")
    wb2.column_dimensions["A"].width = 22; wb2.column_dimensions["B"].width = 70
    _banner(wb2, "Bands (from simulated win%)", 2)
    for label, lo, hi, key in [("Favourable", 55, 101, "fav"), ("Coin-flip / even", 45, 55, "even"),
                               ("Unfavourable", 0, 45, "unfav")]:
        g = [x for x in rows if lo <= x["win"] < hi]
        _row(wb2, [label, ", ".join(f'{x["m"]["faction"]} ({x["win"]}%)' for x in g) or "—"],
             fills=[VF[key], None], bold=True)

    # Tapestry sheet
    wt = wb.create_sheet("Tapestry")
    wt.column_dimensions["A"].width = 42; wt.column_dimensions["B"].width = 95
    _banner(wt, "The rules tapestry (DB/pack-verified)", 2)
    _row(wt, ["Rule", "What it does"], bold=True)
    for n, t in D.RULES:
        _row(wt, [n, t])

    # Profiles sheet
    wp = wb.create_sheet("Profiles")
    for col, w in [("A", 42), ("B", 40), ("C", 55)]:
        wp.column_dimensions[col].width = w
    _banner(wp, "Verified profiles & buffs", 3)
    _row(wp, ["Piece", "Profile / rule", "Note"], bold=True)
    for a, b, c in D.VERIFIED_PROFILES:
        _row(wp, [a, b, c])
    _banner(wp, "Bottom line", 3)
    _row(wp, [D.RECORD_NOTE]); wp.merge_cells(start_row=wp.max_row, start_column=1, end_row=wp.max_row, end_column=3)

    V.stamp_xlsx(wb, "custodes-analysis")
    out = V.out_path("custodes-analysis")
    os.makedirs(os.path.dirname(out), exist_ok=True)
    wb.save(out)
    return out


def runbook(rows, weighted):
    S1 = G.section("Mindset", G.p(G.esc(D.MINDSET)))
    tap = G.table(["Rule", "What it does"], [[G.esc(n), G.esc(t)] for n, t in D.RULES])
    S2 = G.section("Tapestry quick-reference", tap)

    plans = ""
    for r in rows:
        m = r["m"]
        head = (f'{m["faction"]} — <span class="{r["cls"]}">{r["win"]}% ({G.esc(r["verdict"])})</span> '
                f'· you play <b>{G.esc(r["mission"])}</b> vs their {_disp_name(r["disp"])} '
                f'(they play {G.esc(r["opp_mission"])})')
        body = G.p(f'<b>Deciding factor:</b> {G.esc(m["deciding"])}') + G.ul([G.esc(x) for x in m["plan"]])
        if m.get("watch"):
            body += G.p(f'<b>Watch:</b> {G.esc(m["watch"])}')
        plans += G.sub(head, body)
    S3 = G.section("Per-matchup battle plans", plans)

    units = G.table(["Unit (count)", "×", "Role / rules", "Pts"],
                    [[G.esc(n), G.esc(c), G.esc(role), G.esc(pts)] for n, c, role, pts in D.UNITS])
    S4 = G.section("The list", units)
    return [S1, S2, S3, S4]


def main():
    rows, weighted = _rows()
    print("wrote", analysis_xlsx(rows, weighted))            # Analysis = Excel
    print("wrote", G.render("custodes-runbook", runbook(rows, weighted)))  # Runbook = PDF


if __name__ == "__main__":
    main()
