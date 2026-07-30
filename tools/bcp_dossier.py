#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""bcp_dossier.py — full field dossier: YOUR list vs the real event, by archetype.

For each archetype in the event (data/bcp/<event>-archetypes.json), pick a parseable
representative list, sim YOUR army (a rosters handle, default death_rnr) against it, and
weave the sim read together with the hand-written archetype verdict/how-to-play. Writes a
markdown dossier: a MATCHUP MAP (one row per archetype, sorted by field prevalence) plus
full RUNBOOKS for the most common archetypes.

  PYTHONPATH=src python3 tools/bcp_dossier.py [me] [--min 2] [--games 120] [--books 14]
"""
import argparse, html as _html, json, os, subprocess, sys

sys.path.insert(0, os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "src"))
from wh.sim import rosters, bcp, runbook
from wh.sim.runbook import _assess

REC = "data/bcp/lso2026-archetypes.json"

# verdict -> spreadsheet/print fill (the VERDICT is the plan; colour by it, not the directional sim)
VFILL = {"FAVOURABLE": "C7E5CE", "MIRROR": "DBE0E7", "COIN-FLIP": "CBDCEC",
         "UNFAVOURABLE": "F1E3C2", "HARD": "F1CEC9"}
VINK = {"FAVOURABLE": "2f7d52", "MIRROR": "556070", "COIN-FLIP": "35618f",
        "UNFAVOURABLE": "8a6a1a", "HARD": "9e3428"}


def _vkey(verdict):
    return next((k for k in VFILL if (verdict or "").upper().startswith(k)), None)


def _rep_build(arch):
    """A build thunk for the first parseable list in an archetype (its representative)."""
    for p in arch["players"]:
        try:
            b = bcp.builder(p["list_id"])
            b()                      # probe: raises if unparseable / no units / no BSData cut
            return b, p["player"]
        except (Exception, SystemExit):   # skip any rep that won't load; never let it kill the run
            continue
    return None, None


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("me", nargs="?", default="death_rnr")
    ap.add_argument("--min", type=int, default=2, help="min archetype size to include")
    ap.add_argument("--games", type=int, default=120)
    ap.add_argument("--books", type=int, default=14, help="how many top archetypes get a full runbook")
    ap.add_argument("--out", default=None)
    a = ap.parse_args()

    rec = json.load(open(REC, encoding="utf-8"))
    me_builder = getattr(rosters, a.me)
    me_name = me_builder().name
    arches = sorted([(k, v) for k, v in rec["archetypes"].items() if v["size"] >= a.min],
                    key=lambda kv: -kv[1]["size"])

    rows, books = [], []
    for key, arch in arches:
        build_opp, rep = _rep_build(arch)
        if not build_opp:
            rows.append({"key": key, "n": arch["size"], "verdict": arch["verdict"] or "(TBD)",
                         "read": "—", "margin": None, "rep": "(unparseable reps)", "arch": arch})
            continue
        r = runbook.build(me_builder, build_opp, games=a.games)
        head, wincon = _assess(r)
        margin = (r["ctrl"][3] + r["ctrl"][4]) / 2 - (r["octrl"][3] + r["octrl"][4]) / 2
        rows.append({"key": key, "n": arch["size"], "verdict": arch["verdict"] or "(TBD)",
                     "read": head.split(" —")[0].split()[0], "margin": margin, "rep": rep,
                     "arch": arch, "r": r, "s": runbook.structured(r)})
        sys.stderr.write(f"# {key}: {rows[-1]['read']} (rep {rep})\n"); sys.stderr.flush()

    # ---- render markdown ----
    disp = me_builder().disposition
    L = [f"# Field Dossier — {me_name}",
         f"*vs the {rec['event']} field ({rec['n_lists']} lists / {rec['n_archetypes']} archetypes). "
         f"Your disposition: **{disp}**. {a.games} games per matchup.*",
         "",
         "> Sim reads are **directional** (see the sim STATUS) — trust the VERDICT + how-to-play "
         "(hand-verified from the Knights seat) as the plan; the sim read is a mechanistic cross-check.",
         "",
         "## Matchup map",
         "",
         "| N | Archetype | Verdict | Sim | Board |",
         "|--:|-----------|---------|-----|------:|"]
    for x in rows:
        mg = f"{x['margin']:+.2f}" if x["margin"] is not None else "—"
        L.append(f"| {x['n']} | {x['key']} | {x['verdict']} | {x['read']} | {mg} |")
    covered = sum(x["n"] for x in rows)
    L += ["", f"*Covers {covered}/{rec['n_lists']} lists ({100*covered//rec['n_lists']}%). "
          f"Board = your avg objective margin R4-5 (mechanistic).*", "", "---", "", "## Runbooks",
          "*Every simmed archetype, most prevalent first (the field guide reads these; the PDF keeps "
          "full runbooks for the top few + compact cards for the rest).*", ""]
    for x in rows:
        if "r" not in x:
            continue
        L += [f"### {x['key']}  ·  {x['n']} in field  ·  {x['verdict']}", ""]
        if x["arch"]["play"]:
            L += [f"**How to play:** {x['arch']['play']}", ""]
        L += ["```", runbook.report(x["r"]).rstrip(), "```", ""]

    stem = (a.out or f"reports/{a.me}-field-dossier.md").rsplit(".", 1)[0]
    os.makedirs(os.path.dirname(stem) or ".", exist_ok=True)
    open(stem + ".md", "w", encoding="utf-8").write("\n".join(L))
    print(f"\n# dossier -> {stem}.md  ({len(rows)} archetypes, {sum(1 for x in rows if 'r' in x)} simmed)",
          file=sys.stderr)
    simmed = [x for x in rows if "s" in x]
    render_xlsx(simmed, me_name, disp, rec, stem + "-analysis.xlsx")
    render_pdf(simmed, me_name, disp, rec, a.games, stem + ".pdf", a.books)


def render_xlsx(rows, me_name, disp, rec, path):
    """Analysis SPREADSHEET: Matchup Map (one row per archetype) + Threats & Plan (per archetype)."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    wb = Workbook()
    hdr = Font(bold=True, color="FFFFFF", name="Calibri"); hdrfill = PatternFill("solid", fgColor="20242B")
    wrap = Alignment(wrap_text=True, vertical="top")

    def sheet(ws, headers, data, widths, vcol):
        for c, h in enumerate(headers, 1):
            cell = ws.cell(1, c, h); cell.font = hdr; cell.fill = hdrfill; cell.alignment = wrap
        for ri, (verdict, cells) in enumerate(data, 2):
            fill = VFILL.get(_vkey(verdict))
            for c, v in enumerate(cells, 1):
                cell = ws.cell(ri, c, v); cell.alignment = wrap
                if fill and c == vcol:
                    cell.fill = PatternFill("solid", fgColor=fill)
                    cell.font = Font(bold=True, color=VINK.get(_vkey(verdict), "000000"))
        for c, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(c)].width = w
        ws.freeze_panes = "A2"

    ws1 = wb.active; ws1.title = "Matchup Map"
    sheet(ws1, ["# Pilots", "Archetype", "Faction", "Verdict", "Sim (directional)", "Board",
                "Biggest threat", "Top OC holder", "How to play"],
          [(x["verdict"], [x["n"], x["arch"]["detachment"], x["arch"]["faction"],
                           x["verdict"], x["read"], round(x["margin"], 2) if x["margin"] is not None else "",
                           (x["s"]["priority_kills"][0][0] if x["s"]["priority_kills"] else
                            (x["s"]["play_around"][0][0] if x["s"]["play_around"] else "-")),
                           (f"{x['s']['lynchpins'][0][0]} (OC{x['s']['lynchpins'][0][1]})"
                            if x["s"]["lynchpins"] else "-"),
                           x["arch"]["play"]]) for x in rows],
          [8, 30, 20, 30, 12, 7, 22, 22, 80], vcol=4)

    ws2 = wb.create_sheet("Threats & Plan")

    def lst(items, fmt):
        return "\n".join(fmt(i) for i in items) or "-"
    sheet(ws2, ["Archetype", "Verdict", "Priority kills (dmg / removable%)", "Play around (survive% / dmg)",
                "Board lynchpins (OC — who holds objectives)", "Your workhorses", "Your liabilities",
                "Secondaries: lean", "The trap"],
          [(x["verdict"], [x["arch"]["detachment"], x["verdict"],
                           lst(x["s"]["priority_kills"], lambda t: f"{t[0]} — {t[1]}w, {t[2]}% kill"),
                           lst(x["s"]["play_around"], lambda t: f"{t[0]} — {t[1]}% lives, {t[2]}w"),
                           lst(x["s"]["lynchpins"], lambda t: f"{t[0]} — OC{t[1]} ({t[3]})"),
                           lst(x["s"]["workhorses"], lambda t: f"{t[0]} — {t[1]}w"),
                           lst(x["s"]["liabilities"], lambda t: f"{t[0]} — {t[1]}%"),
                           "\n".join(x["s"]["sec_lean"]) or "-", x["s"]["trap"]]) for x in rows],
          [30, 30, 34, 30, 30, 26, 22, 26, 46], vcol=2)
    for r in range(2, ws2.max_row + 1):
        ws2.row_dimensions[r].height = 108
    wb.save(path)
    print(f"# analysis  -> {path}", file=sys.stderr)


_PDF_CSS = """
/* soffice HTML->PDF quirk: a wrapping block WITH a background/full border renders one box PER LINE.
   So NO backgrounds/full borders on multi-line text — verdict colour via TEXT, thin top-rule between
   runbooks, and the (single-line) table cells + pills keep their fills safely. */
body{font-family:'DejaVu Sans',Arial,sans-serif;color:#22262d;max-width:900px;margin:0 auto;
   padding:24px;line-height:1.4;font-size:11px}
.eyebrow{font-family:'DejaVu Sans Mono',monospace;font-size:9px;letter-spacing:2px;color:#9a6f1c;font-weight:bold}
h1{font-size:23px;margin:2px 0 4px;color:#1a1d23;border-bottom:3px solid #c69a3a;padding-bottom:5px}
.sub{color:#5a6069;font-size:11px;margin:2px 0}
.disp{font-family:'DejaVu Sans Mono',monospace;font-size:9px;color:#7a5713;font-weight:bold}
.note{font-size:10px;color:#5a4a12;margin:12px 0;font-style:italic}
h2{font-size:14px;color:#5a4a12;border-bottom:2px solid #c69a3a;padding-bottom:3px;margin:24px 0 10px}
table{border-collapse:collapse;width:100%;font-size:10px;margin:6px 0 14px}
th{background:#20242b;color:#fff;text-align:left;padding:5px 7px;font-size:10px}
td{padding:5px 7px;border-bottom:1px solid #e0dccf;vertical-align:top}
td.n{font-weight:bold;font-size:13px;text-align:center}
.pill{padding:2px 7px;font-weight:bold;font-size:10px;white-space:nowrap}
.book{margin:20px 0 0;page-break-inside:avoid}   /* NO border-top: soffice propagates it to every <pre> line as a hairline rule; the coloured bold title separates books */
.book .bt{font-weight:bold;font-size:14px}
.book .bmeta{font-family:'DejaVu Sans Mono',monospace;font-size:9px;color:#7a7f88;margin:2px 0 6px}
.book .play{font-size:10.5px;color:#22262d;margin:0 0 7px}
.book.compact{margin:9px 0 0}
.book .rb{font-family:'DejaVu Sans Mono',monospace;font-size:9px;white-space:pre-wrap;margin:0;color:#3a3f47}
.book .rb b{color:#111}
.foot{margin-top:16px;font-family:'DejaVu Sans Mono',monospace;font-size:9px;color:#8b909a;
   border-top:1px solid #ddd;padding-top:8px}
"""


def render_pdf(rows, me_name, disp, rec, games, path, books=14):
    e = _html.escape
    import re as _re
    boldrb = lambda s: _re.sub(r"^(READ:|WIN CONDITION:|POSTURE:|DEPLOYMENT:|THE TRAP:|PRIORITY KILLS.*?:|"
                              r"PLAY AROUND.*?:|YOUR WORKHORSES.*?:|YOUR LIABILITIES.*?:|BOARD CONTROL.*?:)",
                              r"<b>\1</b>", e(s), flags=_re.M)
    H = [f"<html><head><meta charset='utf-8'><style>{_PDF_CSS}</style></head><body>",
         "<div class='eyebrow'>Imperial Knights · Field Manual</div>",
         f"<h1>{e(me_name)}</h1>",
         f"<div class='sub'>vs the {e(rec['event'])} field · {rec['n_lists']} lists / {rec['n_archetypes']} "
         f"archetypes · {games} games/matchup</div>",
         f"<div class='sub'>Your locked disposition: <span class='disp'>{e(disp)}</span></div>",
         "<div class='note'><b>Read this first:</b> the VERDICT + how-to-play are hand-set from the Knights "
         "seat — that is the plan. The <b>Sim</b> column is the positional simulator: <b>directional only</b> "
         "(it over-credits you in melee and vs reanimation/hordes), a cross-check, not a prediction.</div>",
         "<h2>Matchup map</h2>",
         "<table><tr><th># Pilots</th><th>Archetype</th><th>Verdict</th><th>Sim</th><th>Board</th>"
         "<th>How to play (short)</th></tr>"]
    for x in rows:
        vk = _vkey(x["verdict"])
        pill = (f"background:#{VFILL.get(vk,'e8e8e8')};color:#{VINK.get(vk,'333')}") if vk else ""
        mg = f"{x['margin']:+.2f}" if x["margin"] is not None else "—"
        short = x["arch"]["play"].split(".")[0] + "." if x["arch"]["play"] else ""
        H.append(f"<tr><td class='n'>{x['n']}</td><td><b>{e(x['arch']['detachment'])}</b><br>"
                 f"<span style='color:#8b909a'>{e(x['arch']['faction'])}</span></td>"
                 f"<td><span class='pill' style='{pill}'>{e(x['verdict'].split(' —')[0].split(' (')[0])}</span></td>"
                 f"<td>{e(x['read'])}</td><td style='font-variant-numeric:tabular-nums'>{mg}</td>"
                 f"<td>{e(short)}</td></tr>")
    H.append(f"</table><h2>Runbooks — top {min(books, len(rows))} archetypes</h2>")
    for x in rows[:books]:
        if not x["r"]:
            continue
        vk = _vkey(x["verdict"]); vc = "#" + (VINK.get(vk, "c69a3a"))
        H.append(f"<div class='book'>"
                 f"<div class='bt' style='color:{vc}'>{e(x['arch']['detachment'])}</div>"
                 f"<div class='bmeta'>{e(x['arch']['faction'])} · {x['n']} pilots · {e(x['verdict'])}</div>"
                 + (f"<div class='play'><b>How to play:</b> {e(x['arch']['play'])}</div>" if x['arch']['play'] else "")
                 + f"<pre class='rb'>{boldrb(runbook.report(x['r']).rstrip())}</pre></div>")
    # the remaining archetypes as compact how-to-play cards (no full runbook — keeps the PDF printable)
    rest = [x for x in rows[books:] if x["arch"]["play"]]
    if rest:
        H.append("<h2>Rest of the field — how-to-play</h2>")
        for x in rest:
            vk = _vkey(x["verdict"]); vc = "#" + (VINK.get(vk, "c69a3a"))
            pill = f"background:#{VFILL.get(vk,'e8e8e8')};color:#{VINK.get(vk,'333')}" if vk else ""
            H.append(f"<div class='book compact'>"
                     f"<div class='bt' style='color:{vc}'>{e(x['arch']['detachment'])} "
                     f"<span class='pill' style='{pill}'>{e(x['verdict'].split(' —')[0].split(' (')[0])}</span></div>"
                     f"<div class='bmeta'>{e(x['arch']['faction'])} · {x['n']} pilots · sim {e(x['read'])}</div>"
                     f"<div class='play'>{e(x['arch']['play'])}</div></div>")
    H.append(f"<div class='foot'>wh.sim positional simulator · mechanistic matchup analysis, not a win-rate "
             f"oracle · {len(rows)} archetypes · field: {e(rec['event'])}.</div></body></html>")
    src = path.rsplit(".", 1)[0] + "_src.html"
    open(src, "w", encoding="utf-8").write("\n".join(H))
    try:
        subprocess.run(["soffice", "--headless", "--convert-to", "pdf", "--outdir",
                        os.path.dirname(path) or ".", src], check=True, capture_output=True, timeout=180)
        made = src.rsplit(".", 1)[0] + ".pdf"
        if os.path.exists(made):
            os.replace(made, path)
        print(f"# pdf       -> {path}", file=sys.stderr)
    except Exception as ex:
        print(f"# pdf render skipped ({ex}); HTML at {src}", file=sys.stderr)
    finally:
        if os.path.exists(src):
            os.remove(src)


if __name__ == "__main__":
    main()
