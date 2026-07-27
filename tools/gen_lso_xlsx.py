#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Generate the LSO Knights list + analysis workbook (Excel) from tools/lso_data.py.
   PYTHONPATH=tools python3 tools/gen_lso_xlsx.py
Output: docs/Reports & Plans/LSO-Knights-List-and-Analysis.xlsx
"""
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import lso_data as D
import doc_versions as V

OUT = V.out_path("lso-analysis")

HEAD = PatternFill("solid", fgColor="1F3864")
SUB = PatternFill("solid", fgColor="D6DCE4")
WHITEB = Font(bold=True, color="FFFFFF")
BOLD = Font(bold=True)
WRAP = Alignment(wrap_text=True, vertical="top")
TOP = Alignment(vertical="top")
CTR = Alignment(horizontal="center", vertical="center")
THIN = Border(*(Side(style="thin", color="BFBFBF"),) * 4)

VERDICT_FILL = {
    "FAVOURABLE": "C6EFCE",
    "COIN-FLIP": "FFEB9C",
    "COIN-FLIP (lean unfavourable)": "FFE0A3",
    "EXPECT-A-LOSS (winnable)": "FCD5B4",
    "UNFAVOURABLE": "F8CBAD",
    "HARD-LOSS": "F4B084",
    "AUTO-LOSS": "FF7C80",
    "PRELIM": "D9D9D9",
}


def vfill(v):
    for k in (v, v.split(" (")[0]):
        if k in VERDICT_FILL:
            return PatternFill("solid", fgColor=VERDICT_FILL[k])
    return PatternFill("solid", fgColor="FFFFFF")


def hrow(ws, r, cols, fill=HEAD, font=WHITEB):
    for c, val in enumerate(cols, 1):
        cell = ws.cell(r, c, val); cell.fill = fill; cell.font = font
        cell.alignment = WRAP; cell.border = THIN


def widths(ws, ws_widths):
    for i, w in enumerate(ws_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def sheet_list(wb):
    ws = wb.active; ws.title = "List"
    ws.merge_cells("A1:E1")
    t = ws.cell(1, 1, f"IMPERIAL KNIGHTS — {D.LIST_NAME}"); t.font = Font(bold=True, size=14, color="FFFFFF"); t.fill = HEAD
    ws.cell(2, 1, "Detachments"); ws.cell(2, 2, D.DETACHMENTS)
    ws.cell(3, 1, "Disposition"); ws.cell(3, 2, D.DISPOSITION).alignment = WRAP
    ws.cell(4, 1, "Total"); ws.cell(4, 2, D.LIST_TOTAL).alignment = WRAP
    for r in (2, 3, 4):
        ws.cell(r, 1).font = BOLD; ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=5)
    ws.cell(6, 1, "Rationale").font = BOLD; ws.merge_cells("B6:E6"); ws.cell(6, 2, D.LIST_RATIONALE).alignment = WRAP
    r = 8
    hrow(ws, r, ["Unit", "x", "Wargear", "~Pts", "Role in the plan"]); r += 1
    for name, cnt, wg, pts, role in D.LIST_UNITS:
        ws.cell(r, 1, name).font = BOLD; ws.cell(r, 2, cnt).alignment = CTR
        ws.cell(r, 3, wg).alignment = WRAP; ws.cell(r, 4, pts).alignment = CTR
        ws.cell(r, 5, role).alignment = WRAP
        for c in range(1, 6): ws.cell(r, c).border = THIN
        r += 1
    r += 1
    ws.cell(r, 1, "ENHANCEMENTS to add (fix the list to 2000; Rotate is free)").font = BOLD
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5); r += 1
    hrow(ws, r, ["Enhancement", "On", "Pts", "Why", ""], fill=SUB, font=BOLD); r += 1
    for name, on, pts, why in D.ENHANCEMENTS:
        ws.cell(r, 1, name).font = BOLD; ws.cell(r, 2, on); ws.cell(r, 3, pts).alignment = CTR
        ws.merge_cells(start_row=r, start_column=4, end_row=r, end_column=5); ws.cell(r, 4, why).alignment = WRAP
        for c in range(1, 6): ws.cell(r, c).border = THIN
        r += 1
    widths(ws, [26, 5, 40, 10, 60])
    for row in range(9, r): ws.row_dimensions[row].height = 42


def sheet_matchups(wb):
    ws = wb.create_sheet("Matchups")
    ws.merge_cells("A1:G1")
    t = ws.cell(1, 1, "VERIFIED MATCHUP MATRIX — all archetypes (11E, 2026-07-26; re-verified for over-optimism)")
    t.font = Font(bold=True, size=12, color="FFFFFF"); t.fill = HEAD
    hrow(ws, 2, ["Faction", "Archetype", "Prev.", "Verdict", "Deciding factor", "The heist (kill-order / positioning)", "Kill-priority"])
    r = 3
    for m in D.MATCHUPS:
        ws.cell(r, 1, m["faction"]).alignment = WRAP; ws.cell(r, 1).font = BOLD
        ws.cell(r, 2, m["archetype"]).alignment = WRAP
        ws.cell(r, 3, m["prev"]).alignment = CTR
        vc = ws.cell(r, 4, m["verdict"]); vc.fill = vfill(m["verdict"]); vc.alignment = WRAP; vc.font = BOLD
        ws.cell(r, 5, m["deciding"]).alignment = WRAP
        ws.cell(r, 6, "\n".join("• " + h for h in m["heist"])).alignment = WRAP
        ws.cell(r, 7, m["kill_priority"]).alignment = WRAP
        for c in range(1, 8): ws.cell(r, c).border = THIN
        ws.row_dimensions[r].height = max(60, 14 * max(len(m["heist"]), 3))
        r += 1
    widths(ws, [16, 26, 6, 18, 34, 52, 26])
    ws.freeze_panes = "A3"


def sheet_rules(wb):
    ws = wb.create_sheet("Rules & Mindset")
    ws.merge_cells("A1:B1"); t = ws.cell(1, 1, "KNIGHT RULES CHEAT-SHEET"); t.font = WHITEB; t.fill = HEAD
    hrow(ws, 2, ["Rule", "What it means"])
    r = 3
    for name, desc in D.RULES:
        ws.cell(r, 1, name).font = BOLD; ws.cell(r, 1).alignment = TOP
        ws.cell(r, 2, desc).alignment = WRAP
        for c in (1, 2): ws.cell(r, c).border = THIN
        ws.row_dimensions[r].height = 40; r += 1
    r += 1
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
    ws.cell(r, 1, "MINDSET").font = WHITEB; ws.cell(r, 1).fill = HEAD; r += 1
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
    ws.cell(r, 1, D.MINDSET).alignment = WRAP; ws.row_dimensions[r].height = 90
    widths(ws, [24, 90])


def sheet_bands(wb):
    ws = wb.create_sheet("Record & Bands")
    ws.merge_cells("A1:B1"); t = ws.cell(1, 1, f"REALISTIC RECORD — {D.EVENT}"); t.font = WHITEB; t.fill = HEAD
    ws.merge_cells("A2:B2"); ws.cell(2, 1, D.RECORD_NOTE).alignment = WRAP; ws.row_dimensions[2].height = 150
    r = 4
    hrow(ws, r, ["Band", "Matchups"]); r += 1
    for band, lists in D.BANDS.items():
        b = ws.cell(r, 1, band); b.font = BOLD; b.fill = vfill(band.split(" (")[0].upper()); b.alignment = WRAP
        ws.cell(r, 2, ", ".join(lists)).alignment = WRAP
        for c in (1, 2): ws.cell(r, c).border = THIN
        ws.row_dimensions[r].height = 34; r += 1
    widths(ws, [30, 84])


def main():
    wb = Workbook()
    sheet_list(wb); sheet_matchups(wb); sheet_rules(wb); sheet_bands(wb)
    os.makedirs(os.path.dirname(OUT), exist_ok=True)
    V.stamp_xlsx(wb, "lso-analysis")
    wb.save(OUT)
    print("wrote", OUT, "-", len(D.MATCHUPS), "matchups")


if __name__ == "__main__":
    main()
