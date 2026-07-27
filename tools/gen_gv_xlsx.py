#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Great Value LSO analysis workbook (Excel) from tools/gv_data.py + mc_gv_sim.py.
   PYTHONPATH=tools:src python3 tools/gen_gv_xlsx.py
Output: docs/Reports & Plans/GV-LSO-Analysis.xlsx
"""
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import gv_data as D
import doc_versions as V

OUT = V.out_path("gv-analysis")
HEAD = PatternFill("solid", fgColor="7F6000")   # IF gold
BOLD = Font(bold=True); WHITEB = Font(bold=True, color="FFFFFF")
WRAP = Alignment(wrap_text=True, vertical="top"); CTR = Alignment(horizontal="center", vertical="center")
THIN = Border(*(Side(style="thin", color="BFBFBF"),) * 4)
VF = {"FAV": "C6EFCE", "COIN": "FFEB9C", "COIN-": "FCD5B4", "HARD": "F4B084"}


def vfill(v):
    for k in ("COIN-", "COIN", "FAV", "HARD"):
        if v.upper().startswith(k):
            return PatternFill("solid", fgColor=VF[k])
    return PatternFill("solid", fgColor="FFFFFF")


def title(ws, r, text, span, size=13):
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=span)
    c = ws.cell(r, 1, text); c.font = Font(bold=True, size=size, color="FFFFFF"); c.fill = HEAD; c.alignment = WRAP


def hrow(ws, r, cols):
    for c, v in enumerate(cols, 1):
        cell = ws.cell(r, c, v); cell.fill = HEAD; cell.font = WHITEB; cell.alignment = WRAP; cell.border = THIN


def widths(ws, ws_w):
    for i, wd in enumerate(ws_w, 1):
        ws.column_dimensions[get_column_letter(i)].width = wd


def sheet_list(wb):
    ws = wb.active; ws.title = "List"
    title(ws, 1, D.LIST_NAME, 4, 14)
    ws.cell(2, 1, "Detachments").font = BOLD; ws.merge_cells("B2:D2"); ws.cell(2, 2, D.DETACHMENTS)
    ws.cell(3, 1, "Disposition").font = BOLD; ws.merge_cells("B3:D3"); ws.cell(3, 2, D.DISPOSITION).alignment = WRAP
    ws.cell(4, 1, "Total").font = BOLD; ws.merge_cells("B4:D4"); ws.cell(4, 2, D.LIST_TOTAL).alignment = WRAP
    ws.cell(6, 1, "Identity").font = BOLD; ws.merge_cells("A6:D6")
    ws.merge_cells("A7:D7"); ws.cell(7, 1, D.IDENTITY).alignment = WRAP; ws.row_dimensions[7].height = 110
    r = 9; hrow(ws, r, ["Unit", "x", "Role", "~Pts"]); r += 1
    for name, cnt, role, pts in D.UNITS:
        ws.cell(r, 1, name).font = BOLD; ws.cell(r, 2, cnt).alignment = CTR
        ws.cell(r, 3, role).alignment = WRAP; ws.cell(r, 4, pts).alignment = CTR
        for c in range(1, 5): ws.cell(r, c).border = THIN
        ws.row_dimensions[r].height = 46; r += 1
    widths(ws, [30, 5, 78, 8])


def sheet_matchups(wb):
    ws = wb.create_sheet("Matchups")
    title(ws, 1, "GREAT VALUE vs the meta (n=70 listhammer, GV's perspective)", 5)
    hrow(ws, 2, ["Faction", "Archetype", "Prev.", "Verdict", "Deciding factor / game plan"])
    r = 3
    for m in D.MATCHUPS:
        ws.cell(r, 1, m["faction"]).font = BOLD; ws.cell(r, 1).alignment = WRAP
        ws.cell(r, 2, m["archetype"]).alignment = WRAP
        ws.cell(r, 3, m["prev"]).alignment = WRAP
        vc = ws.cell(r, 4, m["verdict"]); vc.fill = vfill(m["verdict"]); vc.font = BOLD; vc.alignment = WRAP
        plan = m["deciding"] + "\n\nPLAN: " + "  ".join("• " + p for p in m["plan"]) + "\n\nWATCH: " + m["watch"]
        ws.cell(r, 5, plan).alignment = WRAP
        for c in range(1, 6): ws.cell(r, c).border = THIN
        ws.row_dimensions[r].height = 150; r += 1
    widths(ws, [16, 30, 10, 22, 78]); ws.freeze_panes = "A3"


def sheet_sim(wb):
    ws = wb.create_sheet("Sim Verification")
    title(ws, 1, "DATA-DRIVEN SIM — GV vs the meta (Oath output from data/bsdata via wh.mathhammer)", 6)
    ws.merge_cells("A2:F2"); ws.cell(2, 1, "800 games/archetype. tools/mc_gv_sim.py. Oath dmg = GV's Oath-boosted "
                  "convergence onto the priority target/turn; brick = the Lysander charge. Indicative — coin-flips "
                  "are the practice matchups.").alignment = WRAP; ws.row_dimensions[2].height = 42
    hrow(ws, 3, ["Archetype", "Prev.", "Verdict", "Oath dmg", "Brick melee", "GV win%"])
    r = 4
    try:
        import mc_gv_sim
        rows = mc_gv_sim.results(800)
    except Exception as e:
        rows = []; ws.cell(r, 1, f"(sim unavailable: {e})")
    for x in rows:
        ws.cell(r, 1, x["archetype"]).alignment = WRAP
        ws.cell(r, 2, x["prev"]).alignment = CTR
        vc = ws.cell(r, 3, x["verdict"]); vc.fill = vfill(x["verdict"]); vc.font = BOLD; vc.alignment = CTR
        ws.cell(r, 4, x["oath"]).alignment = CTR; ws.cell(r, 5, x["brick"]).alignment = CTR
        wc = ws.cell(r, 6, f"{x['win']}%"); wc.font = BOLD; wc.alignment = CTR
        for c in range(1, 7): ws.cell(r, c).border = THIN
        r += 1
    if rows:
        prev = sum(x["prev"] for x in rows); wr = sum(x["prev"] * x["win"] for x in rows) / prev
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
        ws.cell(r, 1, f"Prevalence-weighted GV win rate ≈ {wr:.0f}% — strong all-comers; hardest = Orks horde.").font = BOLD
    widths(ws, [34, 7, 8, 10, 12, 9]); ws.freeze_panes = "A4"


def sheet_rules(wb):
    ws = wb.create_sheet("Tapestry & Mindset")
    title(ws, 1, "THE TAPESTRY — how Great Value wins", 2)
    hrow(ws, 2, ["Thread", "What it does"])
    r = 3
    for name, desc in D.RULES:
        ws.cell(r, 1, name).font = BOLD; ws.cell(r, 1).alignment = WRAP
        ws.cell(r, 2, desc).alignment = WRAP
        for c in (1, 2): ws.cell(r, c).border = THIN
        ws.row_dimensions[r].height = 46; r += 1
    r += 1; ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
    ws.cell(r, 1, "MINDSET").font = WHITEB; ws.cell(r, 1).fill = HEAD; r += 1
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
    ws.cell(r, 1, D.MINDSET).alignment = WRAP; ws.row_dimensions[r].height = 90; r += 2
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
    ws.cell(r, 1, "RECORD").font = WHITEB; ws.cell(r, 1).fill = HEAD; r += 1
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
    ws.cell(r, 1, D.RECORD_NOTE).alignment = WRAP; ws.row_dimensions[r].height = 150
    widths(ws, [26, 92])


def main():
    wb = Workbook()
    sheet_list(wb); sheet_matchups(wb); sheet_sim(wb); sheet_rules(wb)
    os.makedirs(os.path.dirname(OUT), exist_ok=True)
    V.stamp_xlsx(wb, "gv-analysis")
    wb.save(OUT)
    print("wrote", OUT, "-", len(D.MATCHUPS), "matchups")


if __name__ == "__main__":
    main()
