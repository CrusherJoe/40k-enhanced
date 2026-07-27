#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""LSO Knights LIST DECISION workbook — 2 Castellan/1 Lancer (A) vs 1 Castellan/2 Lancer (B).
   PYTHONPATH=tools python3 tools/gen_lso_decision_xlsx.py
Output: docs/Reports & Plans/LSO-Knights-List-Decision.xlsx
Grounded in the listhammer top-finishing sample (n=75) + verified 11E profiles.
"""
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import lso_data as D

OUT = "docs/Reports & Plans/LSO-Knights-List-Decision.xlsx"

HEAD = PatternFill("solid", fgColor="1F3864")
SUB = PatternFill("solid", fgColor="D6DCE4")
WHITEB = Font(bold=True, color="FFFFFF")
BOLD = Font(bold=True)
WRAP = Alignment(wrap_text=True, vertical="top")
CTR = Alignment(horizontal="center", vertical="center")
THIN = Border(*(Side(style="thin", color="BFBFBF"),) * 4)

LEAN_FILL = {"A": "BDD7EE", "B": "FCE4D6", "EVEN": "E2E2E2"}
LEAN_FILL2 = {"A (slight)": "DDEBF7", "B (slight)": "FDECE3", "B (weak)": "FDECE3"}


def lfill(v):
    key = v if v in LEAN_FILL else v
    return PatternFill("solid", fgColor=LEAN_FILL.get(v, LEAN_FILL2.get(v, "FFFFFF")))


def title(ws, r, text, span, size=13):
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=span)
    c = ws.cell(r, 1, text); c.font = Font(bold=True, size=size, color="FFFFFF"); c.fill = HEAD; c.alignment = WRAP


def hrow(ws, r, cols, fill=HEAD, font=WHITEB):
    for c, val in enumerate(cols, 1):
        cell = ws.cell(r, c, val); cell.fill = fill; cell.font = font; cell.alignment = WRAP; cell.border = THIN


def widths(ws, ws_widths):
    for i, w in enumerate(ws_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def sheet_decision(wb):
    ws = wb.active; ws.title = "DECISION"
    title(ws, 1, "LSO KNIGHTS — LIST DECISION: 2 Castellan/1 Lancer (A) vs 1 Castellan/2 Lancer (B)", 4, 14)
    ws.merge_cells("A2:D2"); ws.cell(2, 1, D.EVENT + "  ·  " + D.GENERATED).font = Font(italic=True)
    r = 4
    ws.cell(r, 1, "LIST A").font = BOLD; ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=4)
    ws.cell(r, 2, D.LIST_A_NAME + " — " + D.LIST_A_IDENTITY).alignment = WRAP
    ws.cell(r, 1).fill = PatternFill("solid", fgColor="BDD7EE"); ws.row_dimensions[r].height = 60; r += 1
    ws.cell(r, 1, "LIST B").font = BOLD; ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=4)
    ws.cell(r, 2, D.LIST_B_NAME + " — " + D.LIST_B_IDENTITY).alignment = WRAP
    ws.cell(r, 1).fill = PatternFill("solid", fgColor="FCE4D6"); ws.row_dimensions[r].height = 60; r += 2

    # the trade
    ws.cell(r, 1, "THE TRADE").font = WHITEB; ws.cell(r, 1).fill = HEAD
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4); r += 1
    for name, kind, txt in D.SIM_DELTA:
        ws.cell(r, 1, name).font = BOLD; ws.cell(r, 1).alignment = WRAP
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=4)
        ws.cell(r, 2, txt).alignment = WRAP; ws.row_dimensions[r].height = 78; r += 1
    r += 1

    # tally
    ws.cell(r, 1, "PREVALENCE-WEIGHTED TALLY (n=75 top lists)").font = WHITEB; ws.cell(r, 1).fill = HEAD
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4); r += 1
    fills = {"A": "BDD7EE", "B": "FCE4D6", "EVEN": "E2E2E2"}
    for bucket, items in D.DECISION_TALLY.items():
        key = "A" if "LIST A" in bucket else ("B" if "LIST B" in bucket else "EVEN")
        ws.cell(r, 1, bucket).font = BOLD; ws.cell(r, 1).fill = PatternFill("solid", fgColor=fills[key])
        ws.cell(r, 1).alignment = WRAP
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=4)
        ws.cell(r, 2, ", ".join(items)).alignment = WRAP; ws.row_dimensions[r].height = 46; r += 1
    r += 1

    # recommendation
    ws.cell(r, 1, "RECOMMENDATION & REASONING").font = WHITEB; ws.cell(r, 1).fill = HEAD
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4); r += 1
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
    ws.cell(r, 1, D.DECISION).alignment = WRAP; ws.row_dimensions[r].height = 340
    widths(ws, [24, 30, 30, 34])


def sheet_meta(wb):
    ws = wb.create_sheet("Meta (n=70)")
    title(ws, 1, "WINNERS' META — top-finishing lists, ~14 late-July-2026 GTs (n=75)", 5)
    ws.merge_cells("A2:E2"); ws.cell(2, 1, D.OBSERVED_META_NOTE).alignment = WRAP; ws.row_dimensions[2].height = 58
    hrow(ws, 3, ["Faction / archetype", "# of top lists", "Threat character", "Favours", "Note"])
    r = 4
    for name, cnt, char, fav, note in D.META:
        ws.cell(r, 1, name).font = BOLD; ws.cell(r, 1).alignment = WRAP
        ws.cell(r, 2, cnt).alignment = CTR
        ws.cell(r, 3, char).alignment = WRAP
        fc = ws.cell(r, 4, fav); fc.fill = lfill(fav.split()[0] if fav.split() else fav); fc.alignment = CTR; fc.font = BOLD
        ws.cell(r, 5, note).alignment = WRAP
        for c in range(1, 6): ws.cell(r, c).border = THIN
        ws.row_dimensions[r].height = 44; r += 1
    widths(ws, [40, 12, 30, 12, 52]); ws.freeze_panes = "A4"


def sheet_list(wb, name, sheetname, identity, units, fill):
    ws = wb.create_sheet(sheetname)
    title(ws, 1, name, 5)
    ws.merge_cells("A2:E2"); ws.cell(2, 1, identity).alignment = WRAP; ws.row_dimensions[2].height = 58
    ws.cell(2, 1).fill = PatternFill("solid", fgColor=fill)
    hrow(ws, 3, ["Unit", "x", "Wargear", "~Pts", "Role in the plan"])
    r = 4
    for uname, cnt, wg, pts, role in units:
        ws.cell(r, 1, uname).font = BOLD; ws.cell(r, 2, cnt).alignment = CTR
        ws.cell(r, 3, wg).alignment = WRAP; ws.cell(r, 4, pts).alignment = CTR
        ws.cell(r, 5, role).alignment = WRAP
        for c in range(1, 6): ws.cell(r, c).border = THIN
        ws.row_dimensions[r].height = 54; r += 1
    r += 1
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
    ws.cell(r, 1, D.LIST_TOTAL).alignment = WRAP; ws.row_dimensions[r].height = 46; r += 1
    ws.cell(r, 1, "Enhancements to reach 2000 (Rotate is free):").font = BOLD; r += 1
    for ename, on, pts, why in D.ENHANCEMENTS:
        ws.cell(r, 1, ename).font = BOLD; ws.cell(r, 2, on)
        ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=5); ws.cell(r, 3, f"({pts}) {why}").alignment = WRAP
        ws.row_dimensions[r].height = 30; r += 1
    widths(ws, [24, 5, 46, 10, 52])


def sheet_matchups(wb):
    ws = wb.create_sheet("Matchups A vs B")
    title(ws, 1, "PER-MATCHUP: which list performs better (verified 11E, observed meta)", 6)
    hrow(ws, 2, ["Faction", "Archetype", "Prev.", "Verdict (vs the list generally)", "Better list", "Why that list"])
    r = 3
    for m in D.MATCHUPS:
        lean, why = D.MATCHUP_LEANS[m["key"]]
        ws.cell(r, 1, m["faction"]).font = BOLD; ws.cell(r, 1).alignment = WRAP
        ws.cell(r, 2, m["archetype"]).alignment = WRAP
        ws.cell(r, 3, m["prev"]).alignment = CTR
        ws.cell(r, 4, m["verdict"]).alignment = WRAP
        lc = ws.cell(r, 5, lean); lc.fill = lfill(lean.split()[0]); lc.alignment = CTR; lc.font = BOLD
        ws.cell(r, 6, why).alignment = WRAP
        for c in range(1, 7): ws.cell(r, c).border = THIN
        ws.row_dimensions[r].height = 58; r += 1
    widths(ws, [16, 30, 7, 22, 12, 54]); ws.freeze_panes = "A3"


def sheet_profiles(wb):
    ws = wb.create_sheet("Verified Profiles")
    title(ws, 1, "VERIFIED 11E PROFILES (BSData + faction packs) — the numbers behind the sim", 3)
    hrow(ws, 2, ["Unit / weapon", "Profile", "Note"])
    r = 3
    for wname, prof, note in D.VERIFIED_PROFILES:
        b = wname.strip().startswith("---") or not wname.startswith("  ")
        ws.cell(r, 1, wname).font = BOLD if b else Font()
        if wname.strip().startswith("---"):
            for c in range(1, 4): ws.cell(r, c).fill = SUB
        ws.cell(r, 1).alignment = WRAP
        ws.cell(r, 2, prof).alignment = WRAP; ws.cell(r, 3, note).alignment = WRAP
        for c in range(1, 4): ws.cell(r, c).border = THIN
        r += 1
    widths(ws, [34, 44, 44])


def sheet_sim(wb):
    ws = wb.create_sheet("Sim Verification")
    title(ws, 1, "DATA-DRIVEN SIM — A vs B (my Knights' output computed from data/bsdata via wh.mathhammer)", 9)
    ws.merge_cells("A2:I2")
    ws.cell(2, 1, "800 games/(list x archetype). tools/mc_db_sim.py. akR/akM = enemy anti-Knight EV/turn "
                  "(ranged/melee, DB where present else verified floor); rmv = my EV onto their priority "
                  "target/turn. Confirms: A wins SHOOTING metas, B wins MELEE metas.").alignment = WRAP
    ws.row_dimensions[2].height = 42
    hrow(ws, 3, ["Archetype", "prev", "akR", "akM", "A rmv", "B rmv", "A win%", "B win%", "Better list"])
    r = 4
    try:
        import mc_db_sim
        rows = mc_db_sim.results(800)
    except Exception as e:
        ws.cell(r, 1, f"(sim unavailable: {e})"); rows = []
    for x in rows:
        ws.cell(r, 1, x["archetype"]).alignment = WRAP
        for c, k in enumerate(("prev", "akR", "akM", "remA", "remB", "winA", "winB"), 2):
            ws.cell(r, c, x[k]).alignment = CTR
        bc = ws.cell(r, 9, x["best"]); bc.alignment = CTR; bc.font = BOLD
        bc.fill = lfill(x["best"].replace("~", "").upper()[:1]) if x["best"] != "~even" else lfill("EVEN")
        for c in range(1, 10):
            ws.cell(r, c).border = THIN
        r += 1
    widths(ws, [34, 6, 6, 6, 7, 7, 8, 8, 12])
    ws.freeze_panes = "A4"


def main():
    wb = Workbook()
    sheet_decision(wb)
    sheet_meta(wb)
    sheet_list(wb, D.LIST_A_NAME, "List A", D.LIST_A_IDENTITY, D.LIST_A_UNITS, "BDD7EE")
    sheet_list(wb, D.LIST_B_NAME, "List B", D.LIST_B_IDENTITY, D.LIST_B_UNITS, "FCE4D6")
    sheet_matchups(wb)
    sheet_sim(wb)
    sheet_profiles(wb)
    os.makedirs(os.path.dirname(OUT), exist_ok=True)
    wb.save(OUT)
    print("wrote", OUT, "-", len(D.MATCHUPS), "matchups,", len(D.META), "meta rows")


if __name__ == "__main__":
    main()
