# -*- coding: utf-8 -*-
"""gen_list.py — the GENERIC new-standard deliverable generator: Analysis (Excel) + Runbook (PDF)
for any list's data module + real-mission sim rows. This is the reusable back-end of the standard
"analyze this list" workflow (see the wh-list-analysis-workflow memory).

  build(D, rows, akey, rkey, my_disp, finding_text)
    D    = the list's data module (LIST_NAME, DETACHMENTS, DISPOSITION, RULES, MINDSET, UNITS,
           VERIFIED_PROFILES, RECORD_NOTE, MATCHUPS[key/faction/archetype/deciding/(plan|heist...)]).
    rows = sim results aligned to D.MATCHUPS order, each dict(win, prev, disp, mission, opp_mission).
    akey/rkey = doc_versions keys for the Analysis (.xlsx) and Runbook (.pdf).
Verdicts are DERIVED from win%. Analysis = Excel, Runbook = PDF (no Word), versioned via doc_versions.
"""
import os
import gen_pdf as G
import doc_versions as V
import sim_game
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

WRAP = Alignment(wrap_text=True, vertical="top")
VF = {"fav": "C6EFCE", "even": "FFEB9C", "unfav": "FFC7CE"}


def verdict(w):
    if w >= 62:
        return "Favourable", "fav"
    if w >= 55:
        return "Lean favourable", "fav"
    if w >= 45:
        return "Coin-flip", "even"
    if w >= 40:
        return "Lean unfavourable", "unfav"
    return "Unfavourable", "unfav"


def _dn(k):
    return sim_game.DISP_NAME.get(k, k)


def _merge_rows(D, rows):
    out, tot, wsum = [], 0, 0
    for m, r in zip(D.MATCHUPS, rows):
        vd, cls = verdict(r["win"])
        tot += r["prev"]; wsum += r["prev"] * r["win"]
        out.append(dict(m=m, win=r["win"], disp=r["disp"], mission=r["mission"],
                        opp_mission=r["opp_mission"], verdict=vd, cls=cls))
    return out, round(wsum / tot)


def _bx(ws, text, ncol, HEAD, size=11):
    r = ws.max_row + (1 if ws.max_row > 1 else 0) + 1
    c = ws.cell(r, 1, text); c.font = Font(bold=True, size=size, color="FFFFFF"); c.fill = HEAD; c.alignment = WRAP
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=ncol)


def _rw(ws, cells, fills=None, bold=False):
    r = ws.max_row + 1
    for i, val in enumerate(cells, 1):
        c = ws.cell(r, i, val); c.alignment = WRAP
        if bold:
            c.font = Font(bold=True)
        if fills and i - 1 < len(fills) and fills[i - 1]:
            c.fill = PatternFill("solid", fgColor=fills[i - 1])


def _analysis(D, rows, weighted, akey, my_disp, finding, colour):
    HEAD = PatternFill("solid", fgColor=colour)
    wb = Workbook(); ws = wb.active; ws.title = "Matchups"
    ws.column_dimensions["A"].width = 46
    for col, w in [("B", 16), ("C", 20), ("D", 8), ("E", 16)]:
        ws.column_dimensions[col].width = w
    _bx(ws, getattr(D, "LIST_NAME", "List"), 5, HEAD, 12)
    for line in [getattr(D, "DETACHMENTS", ""), getattr(D, "DISPOSITION", ""),
                 f"Prevalence-weighted win rate: ~{weighted}%.  {finding}"]:
        if line:
            _rw(ws, [line], bold=True)
            ws.merge_cells(start_row=ws.max_row, start_column=1, end_row=ws.max_row, end_column=5)
    _bx(ws, f"Matchups (you = {_dn(my_disp)})", 5, HEAD)
    _rw(ws, ["Faction / archetype", "Opp disposition", "You play", "Win%", "Read"], bold=True)
    for cell in ws[ws.max_row]:
        cell.fill = HEAD; cell.font = Font(bold=True, color="FFFFFF")
    for r in rows:
        m = r["m"]; fill = VF.get(r["cls"])
        _rw(ws, [f'{m["faction"]} — {m["archetype"]}', _dn(r["disp"]), r["mission"], f'{r["win"]}%', r["verdict"]],
            fills=[None, None, None, fill, fill])

    wb2 = wb.create_sheet("Bands")
    wb2.column_dimensions["A"].width = 22; wb2.column_dimensions["B"].width = 92
    _bx(wb2, "Bands (from simulated win%)", 2, HEAD)
    for label, lo, hi, key in [("Favourable 55%+", 55, 101, "fav"), ("Coin-flip 45-54%", 45, 55, "even"),
                               ("Unfavourable <45%", 0, 45, "unfav")]:
        g = sorted([x for x in rows if lo <= x["win"] < hi], key=lambda x: -x["win"])
        _rw(wb2, [label, ", ".join(f'{x["m"]["faction"]} {x["win"]}%' for x in g) or "—"],
            fills=[VF[key], None], bold=True)

    wt = wb.create_sheet("Tapestry")
    wt.column_dimensions["A"].width = 32; wt.column_dimensions["B"].width = 100
    _bx(wt, "The rules tapestry (DB-verified)", 2, HEAD)
    _rw(wt, ["Rule", "What it does"], bold=True)
    for n, t in getattr(D, "RULES", []):
        _rw(wt, [n, t])

    wp = wb.create_sheet("Profiles")
    for col, w in [("A", 40), ("B", 42), ("C", 52)]:
        wp.column_dimensions[col].width = w
    _bx(wp, "Verified profiles", 3, HEAD)
    for row in getattr(D, "VERIFIED_PROFILES", []):
        _rw(wp, [str(x) for x in row])
    if getattr(D, "RECORD_NOTE", ""):
        _bx(wp, "Bottom line", 3, HEAD)
        _rw(wp, [D.RECORD_NOTE]); wp.merge_cells(start_row=wp.max_row, start_column=1, end_row=wp.max_row, end_column=3)

    V.stamp_xlsx(wb, akey)
    out = V.out_path(akey); os.makedirs(os.path.dirname(out), exist_ok=True); wb.save(out)
    return out


# per-matchup narrative fields to render (label, key) — covers both the plan/watch and heist styles
_NARR = [("Plan", "plan"), ("Mission / heist", "heist"), ("Kill priority", "kill_priority"),
         ("Deploy", "deploy"), ("Watch", "watch")]


def _runbook(D, rows, weighted, rkey, my_disp, finding):
    S1 = G.section("Mindset", G.p(G.esc(getattr(D, "MINDSET", ""))) + G.finding(f"Weighted ~{weighted}%. {finding}"))
    S2 = G.section("Tapestry quick-reference",
                   G.table(["Rule", "What it does"], [[G.esc(n), G.esc(t)] for n, t in getattr(D, "RULES", [])]))
    plans = ""
    for r in rows:
        m = r["m"]
        head = (f'{m["faction"]} — <span class="{r["cls"]}">{r["win"]}% ({G.esc(r["verdict"])})</span> · '
                f'you play <b>{G.esc(r["mission"])}</b> vs their {_dn(r["disp"])} (they play {G.esc(r["opp_mission"])})')
        body = G.p(f'<b>Deciding factor:</b> {G.esc(m.get("deciding",""))}')
        for label, key in _NARR:
            val = m.get(key)
            if isinstance(val, (list, tuple)):
                body += G.p(f"<b>{label}:</b>") + G.ul([G.esc(x) for x in val])
            elif val:
                body += G.p(f"<b>{label}:</b> {G.esc(val)}")
        plans += G.sub(head, body)
    S3 = G.section(f"Per-matchup battle plans ({len(rows)} archetypes)", plans)
    units = getattr(D, "UNITS", [])
    ncol = max((len(u) for u in units), default=0)
    hdr = ["Unit", "×", "Role / wargear", "Pts", "Note"][:ncol] or ["Unit"]
    S4 = G.section("The list", G.table(hdr, [[G.esc(x) for x in u] for u in units]))
    return [S1, S2, S3, S4]


def build(D, rows, akey, rkey, my_disp, finding, colour="333333"):
    merged, weighted = _merge_rows(D, rows)
    a = _analysis(D, merged, weighted, akey, my_disp, finding, colour)
    b = G.render(rkey, _runbook(D, merged, weighted, rkey, my_disp, finding))
    return a, b, weighted
