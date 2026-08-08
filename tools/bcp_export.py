#!/usr/bin/env python3
"""bcp_export.py — export the weekly-meta results to a versioned EXCEL workbook + PDF, sorted by faction.

Same numbers as the dashboard (tools/bcp_dashboard.build_data): win rate + field share + top-cut share by
faction / disposition / detachment. Emits:
  docs/reports/meta/40k-11e-Meta-v<MAJOR.MINOR>.xlsx   (3 sheets: Factions, Dispositions, Detachments)
  docs/reports/meta/40k-11e-Meta-v<MAJOR.MINOR>.pdf    (cover + the three tables)

VERSIONING: MAJOR.MINOR in docs/reports/meta/VERSION. The MINOR auto-bumps whenever the data SNAPSHOT changes
(event set / week range / row counts) so each distinct pull gets a new version + a new versioned file (old
revisions stay on disk); a rebuild on unchanged data reproduces the same version. Bump MAJOR by hand for a
structural change to the report itself.  Run:  PYTHONPATH=src python3 tools/bcp_export.py
"""
import sys, os, json, collections, subprocess, datetime as dt
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
import bcp_dashboard as DASH
import bcp_meta as BM
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.properties import PageSetupProperties

OUTDIR = "docs/reports/meta"
MAJOR = 1
TITLE = "Warhammer 40,000 · 11th Edition — Competitive Meta"


def detach_faction_map():
    """Each detachment -> the faction that most commonly runs it (detachments are ~faction-exclusive)."""
    byd = collections.defaultdict(collections.Counter)
    for r in BM.load_corpus():
        if r.get("detach"):
            byd[r["detach"]][r["faction"]] += 1
    return {d: c.most_common(1)[0][0] for d, c in byd.items()}


def rows_for(dim_rows, totals_key="total"):
    """Flatten dashboard dim rows into export rows: name, lists, field%, top%, delta_pp, games, win%."""
    tp = sum(r[totals_key]["players"] for r in dim_rows) or 1
    tt = sum(r[totals_key]["top"] for r in dim_rows) or 1
    out = []
    for r in dim_rows:
        t = r[totals_key]
        field = t["players"] / tp
        top = t["top"] / tt if tt else 0
        win = t["wins"] / t["games"] if t["games"] else None
        out.append(dict(name=r["name"], lists=t["players"], field=field, top=top,
                        delta=top - field, games=t["games"], win=win))
    return out


def snapshot_sig(data):
    s = data["summary"]
    return f"{s['events']}|{s['games']}|{'-'.join(s['week_range'])}|{len(data['faction'])}|{len(data['detachment'])}"


def resolve_version(sig):
    os.makedirs(OUTDIR, exist_ok=True)
    vf = os.path.join(OUTDIR, "VERSION")
    cur_major, cur_minor, cur_sig = MAJOR, 0, ""
    if os.path.exists(vf):
        try:
            ver, cur_sig = open(vf).read().strip().split("|", 1)
            cur_major, cur_minor = (int(x) for x in ver.split("."))
        except Exception:
            pass
    if cur_major != MAJOR:                 # structural bump resets the minor
        cur_minor, cur_sig = 0, ""
    minor = cur_minor if sig == cur_sig else cur_minor + 1
    ver = f"{MAJOR}.{minor}"
    open(vf, "w").write(f"{ver}|{sig}")
    return ver


# ------------------------------------------------------------------ Excel
_HEAD = Font(bold=True, color="FFFFFF"); _HFILL = PatternFill("solid", fgColor="1F3864")
_TITLE = Font(bold=True, size=14); _SUB = Font(size=9, italic=True, color="666666")
_THIN = Border(bottom=Side(style="thin", color="D9D9D9"))


def _sheet(ws, headers, rows, subtitle):
    ws.append([TITLE]); ws["A1"].font = _TITLE
    ws.append([subtitle]); ws["A2"].font = _SUB
    ws.append([])
    hr = ws.max_row + 1
    ws.append(headers)
    for c in ws[hr]:
        c.font = _HEAD; c.fill = _HFILL; c.alignment = Alignment(horizontal="center")
    for row in rows:
        ws.append(row)
    for r in range(hr + 1, ws.max_row + 1):
        for c in ws[r]:
            c.border = _THIN
    ws.freeze_panes = ws.cell(row=hr + 1, column=1)
    # page setup so the xlsx->pdf render is clean: landscape, fit all columns to one page wide, repeat header
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1; ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
    ws.print_title_rows = f"{hr}:{hr}"
    return hr


def _pct(ws, col, r0, r1, fmt="0.0%"):
    for r in range(r0, r1 + 1):
        ws[f"{col}{r}"].number_format = fmt


def write_excel(path, ver, data, dmap, subtitle):
    wb = openpyxl.Workbook()
    # --- Factions (sorted A→Z by faction) ---
    ws = wb.active; ws.title = "Factions"
    fr = sorted(rows_for(data["faction"]), key=lambda x: x["name"].lower())
    hdr = ["Faction", "Lists", "Field share", "Top-10 share", "Δ (top−field)", "Games", "Win rate"]
    body = [[x["name"], x["lists"], x["field"], x["top"], x["delta"], x["games"],
             x["win"] if x["win"] is not None else "—"] for x in fr]
    h = _sheet(ws, hdr, body, subtitle)
    _pct(ws, "C", h + 1, ws.max_row); _pct(ws, "D", h + 1, ws.max_row)
    _pct(ws, "E", h + 1, ws.max_row, "+0.0%;-0.0%"); _pct(ws, "G", h + 1, ws.max_row, "0%")
    for col, w in zip("ABCDEFG", (26, 7, 12, 13, 14, 8, 10)):
        ws.column_dimensions[col].width = w
    # --- Dispositions (sorted by win rate) ---
    ws2 = wb.create_sheet("Dispositions")
    dr = sorted(rows_for(data["disposition"]), key=lambda x: -(x["win"] or 0))
    body = [[x["name"], x["lists"], x["field"], x["top"], x["delta"], x["games"], x["win"]] for x in dr]
    h = _sheet(ws2, hdr[:1] + hdr[1:], body, subtitle)  # same headers, first col relabelled below
    ws2["A4"] = "Disposition"; ws2["A4"].font = _HEAD; ws2["A4"].fill = _HFILL
    _pct(ws2, "C", h + 1, ws2.max_row); _pct(ws2, "D", h + 1, ws2.max_row)
    _pct(ws2, "E", h + 1, ws2.max_row, "+0.0%;-0.0%"); _pct(ws2, "G", h + 1, ws2.max_row, "0%")
    for col, w in zip("ABCDEFG", (18, 7, 12, 13, 14, 8, 10)):
        ws2.column_dimensions[col].width = w
    # --- Detachments (grouped/sorted by FACTION, then win rate; drop tiny n) ---
    ws3 = wb.create_sheet("Detachments")
    det = [x for x in rows_for(data["detachment"]) if x["lists"] >= 5]
    for x in det:
        x["faction"] = dmap.get(x["name"], "—")
    det.sort(key=lambda x: (x["faction"].lower(), -(x["win"] or 0)))
    hdr3 = ["Faction", "Detachment", "Lists", "Field share", "Top-10 share", "Δ (top−field)", "Games", "Win rate"]
    body = [[x["faction"], x["name"], x["lists"], x["field"], x["top"], x["delta"], x["games"], x["win"]] for x in det]
    h = _sheet(ws3, hdr3, body, subtitle + "  ·  detachments with ≥5 lists")
    _pct(ws3, "D", h + 1, ws3.max_row); _pct(ws3, "E", h + 1, ws3.max_row)
    _pct(ws3, "F", h + 1, ws3.max_row, "+0.0%;-0.0%"); _pct(ws3, "H", h + 1, ws3.max_row, "0%")
    for col, w in zip("ABCDEFGH", (22, 30, 7, 12, 13, 14, 8, 10)):
        ws3.column_dimensions[col].width = w
    wb.save(path)


# ------------------------------------------------------------------ PDF: render the XLSX via LibreOffice Calc.
# (LibreOffice's HTML import ignores CSS white-space:nowrap and char-wraps every cell into a sliver; Calc
# renders the sheets at their real column widths with the page-setup above — clean, landscape, fit-to-width.)
def write_pdf(pdf_path, xlsx_path):
    outdir = os.path.dirname(pdf_path)
    subprocess.run(["soffice", "--headless", "--convert-to", "pdf", "--outdir", outdir, xlsx_path],
                   check=True, stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)
    produced = os.path.join(outdir, os.path.splitext(os.path.basename(xlsx_path))[0] + ".pdf")
    if produced != pdf_path and os.path.exists(produced):
        os.replace(produced, pdf_path)
    return pdf_path


def main():
    data = DASH.build_data()
    dmap = detach_faction_map()
    s = data["summary"]
    ver = resolve_version(snapshot_sig(data))
    wr = s["week_range"]
    subtitle = (f"generated {dt.date.today().isoformat()} · {s['events']} GTs, {s['games']:,} games, "
                f"weeks {wr[0]}–{wr[1]}" if wr else f"generated {dt.date.today().isoformat()}")
    os.makedirs(OUTDIR, exist_ok=True)
    xlsx = os.path.join(OUTDIR, f"40k-11e-Meta-v{ver}.xlsx")
    pdf = os.path.join(OUTDIR, f"40k-11e-Meta-v{ver}.pdf")
    write_excel(xlsx, ver, data, dmap, subtitle)
    write_pdf(pdf, xlsx)
    print(f"# v{ver} · {subtitle}")
    print(f"# wrote {xlsx}")
    print(f"# wrote {pdf}")


if __name__ == "__main__":
    main()
