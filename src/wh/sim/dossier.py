"""The DOSSIER — the full tournament-prep deliverable for a list. Assembles, into one document:
  1. YOUR TAPESTRY   — army rule + every chosen detachment's stratagems + the rule->engine mapping.
  2. MATCHUP MAP     — one row per meta archetype: the read, your posture, the single biggest threat,
                       and the one-line game plan. The "how does my list do into X" overview.
  3. RUNBOOKS        — the full per-archetype play guide (priority kills / play-around / workhorses /
                       liabilities / stratagems / the trap).
  4. LIST FIXES      — the optimizer's TESTED swap + detachment recommendations on a chosen matchup.

Mechanistic, not a win% oracle (see the sim STATUS). Writes reports/dossier-<me>.md.

  python -m wh.sim.dossier custodes [--games 250] [--opt necrons]
"""
from __future__ import annotations

import os, subprocess, html as _html

from . import runbook, optimize, rosters, strategy as St
from .gauntlet import tapestry, OPPONENTS, ME_META
import db

_READ_CLASS = {"FAVOURED": "fav", "EVEN": "even", "ATTRITION": "attr", "HARD": "hard"}


def _read_key(read):
    for k in _READ_CLASS:
        if read.startswith(k):
            return _READ_CLASS[k]
    return "even"


def _read_short(head):
    return head.split("—")[0].strip()


def _matchup_row(me_name, opp, r):
    arch = St.archetype(getattr(rosters, opp)())
    head, wincon = runbook._assess(r)
    pk = runbook._priority_kills(r)
    pa = runbook._play_around(r)
    threat = pk[0]["name"] if pk else (pa[0]["name"] if pa else "-")
    posture = r["my_strategy"].name if r["my_strategy"] else "balanced"
    return dict(opp=opp, arch=arch, read=_read_short(head), threat=threat, posture=posture, wincon=wincon)


def gather(me_name="custodes", games=250, opt_opp="necrons", seed=11):
    """Run the sim ONCE and collect the structured pieces (tapestry, map rows, runbooks, fixes)."""
    me_builder = getattr(rosters, me_name)
    slug, dets = ME_META.get(me_name, (getattr(me_builder(), "slug", me_name), ("",)))
    tap = tapestry(me_builder(), slug, dets[0])
    if len(dets) > 1:
        try:
            extra = db.strats(slug).get(dets[1], {})
            tap += f"\n\n## {dets[1]} stratagems ({len(extra)}) — DB\n" + "\n".join(
                f"- **{nm}** ({s.get('cp','?')}CP) — {' '.join(str(s.get('effect','')).split())[:150]}"
                for nm, s in extra.items())
        except Exception:
            pass
    rows, books, details = [], [], []
    for opp in OPPONENTS:
        r = runbook.build(me_builder, getattr(rosters, opp), games=games, seed=seed)
        row = _matchup_row(me_name, opp, r)
        row["arch"] = St.archetype(getattr(rosters, opp)())
        rows.append(row)
        books.append(runbook.report(r))
        s = runbook.structured(r); s["opp_key"] = opp; s["arch"] = row["arch"]
        details.append(s)
    opt = optimize.optimize(me_builder, getattr(rosters, opt_opp),
                            screen=max(400, games // 2), final=games, seed=seed)
    fixes = optimize.report(opt)
    return dict(name=me_builder().name, me=me_name, tap=tap, rows=rows, books=books, details=details,
                fixes=fixes, opt_opp=opt_opp, games=games)


def render_markdown(P):
    out = [f"# TOURNAMENT DOSSIER — {P['name']}", "",
           "Your list is LOCKED at the event. The RUNBOOKS below are your IN-EVENT tool — how to pilot "
           "THIS list through each matchup. The MATCHUP MAP is the cheat sheet. LIST-BUILDING notes at the "
           "end are for choosing your list BETWEEN events, not for mid-tournament.", "",
           P["tap"], "",
           "# MATCHUP MAP", "", f"  {'ARCHETYPE (list)':28} {'READ':18} {'POSTURE':9} BIGGEST THREAT"]
    for row in P["rows"]:
        out.append(f"  {row['opp']+' ('+row['arch']+')':28} {row['read']:18} {row['posture']:9} {row['threat']}")
    out += ["", "# RUNBOOKS (per archetype) — your in-event pilot guide", ""]
    for b in P["books"]:
        out += ["```", b, "```", ""]
    out += ["# LIST-BUILDING NOTES (between events — the list is locked at the tournament)", "",
            f"(optimizer vs {P['opt_opp']}; use this when choosing what to bring, not mid-event)", "```",
            P["fixes"], "```"]
    return "\n".join(out)


_CSS = """
body{font-family:'DejaVu Sans',Arial,sans-serif;color:#1a1a1a;max-width:920px;margin:0 auto;padding:24px;line-height:1.4}
h1{font-size:22px;border-bottom:3px solid #8a6d1a;padding-bottom:6px}
h2{font-size:15px;color:#5a4a12;margin-top:22px}
.sub{color:#666;font-size:12px;margin:2px 0 16px}
table.map{border-collapse:collapse;width:100%;font-size:13px;margin:8px 0 18px}
table.map th{background:#2a2a2a;color:#fff;text-align:left;padding:6px 9px}
table.map td{padding:5px 9px;border-bottom:1px solid #ddd}
.pill{padding:2px 8px;border-radius:10px;font-weight:bold;font-size:11px;color:#fff;white-space:nowrap}
.fav{background:#1f8a3b}.even{background:#2a6db0}.attr{background:#b8860b}.hard{background:#b23030}
.book{border:1px solid #ccc;border-left:4px solid #8a6d1a;border-radius:4px;padding:10px 14px;margin:12px 0;
 white-space:pre-wrap;font-family:'DejaVu Sans Mono',monospace;font-size:11px;page-break-inside:avoid}
.book .rb-title{font-weight:bold;font-size:13px;color:#5a4a12}
pre{white-space:pre-wrap;font-family:'DejaVu Sans Mono',monospace;font-size:11px;background:#f6f4ee;padding:10px;border-radius:4px}
.note{background:#fff6e0;border:1px solid #e0c060;padding:8px 12px;border-radius:4px;font-size:12px;margin:10px 0}
"""


def render_html(P):
    e = _html.escape
    h = [f"<html><head><meta charset='utf-8'><title>Dossier — {e(P['name'])}</title><style>{_CSS}</style></head><body>",
         f"<h1>Tournament Dossier — {e(P['name'])}</h1>",
         f"<div class='sub'>{len(P['rows'])} archetypes · {P['games']} games each · mechanistic matchup analysis "
         "(not a win% prediction — the values are the sim's internal scale)</div>",
         "<div class='note'>Your list is LOCKED at the event. The RUNBOOKS are your IN-EVENT tool — how to pilot "
         "THIS list through each matchup (who to kill, what you can't remove, what to protect, the trap). The MAP is "
         "the cheat sheet. LIST-BUILDING notes at the end are for BETWEEN events, not mid-tournament.</div>",
         "<h2>Matchup Map</h2>",
         "<table class='map'><tr><th>Archetype (list)</th><th>Read</th><th>Posture</th><th>Biggest threat</th></tr>"]
    for row in P["rows"]:
        h.append(f"<tr><td>{e(row['opp'])} <span style='color:#888'>({e(row['arch'])})</span></td>"
                 f"<td><span class='pill {_read_key(row['read'])}'>{e(row['read'])}</span></td>"
                 f"<td>{e(row['posture'])}</td><td>{e(row['threat'])}</td></tr>")
    h.append("</table>")
    # tapestry
    h.append("<h2>Your Tapestry</h2><pre>" + e(P["tap"]) + "</pre>")
    # runbooks
    h.append("<h2>Runbooks (per archetype)</h2>")
    for b in P["books"]:
        lines = b.split("\n")
        title = e(lines[0]); rest = e("\n".join(lines[1:]))
        h.append(f"<div class='book'><span class='rb-title'>{title}</span>\n{rest}</div>")
    h.append("<h2>List-Building Notes <span style='font-weight:400;font-size:12px;color:#888'>"
             "(between events — the list is locked at the tournament)</span></h2><pre>" + e(P["fixes"]) + "</pre>")
    h.append("</body></html>")
    return "\n".join(h)


_READ_FILL = {"FAVOURED": "C7E5CE", "EVEN": "CBDCEC", "ATTRITION": "F0E2C0", "HARD": "F0CCC7"}


def render_xlsx(P, path):
    """Analysis SPREADSHEET (.xlsx): a Matchup Map sheet + a per-matchup Threats & Plan sheet."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    wb = Workbook()
    hdr = Font(bold=True, color="FFFFFF"); hdrfill = PatternFill("solid", fgColor="2A2A2A")
    wrap = Alignment(wrap_text=True, vertical="top")

    def sheet(ws, headers, rows, widths, fills=None):
        for c, h in enumerate(headers, 1):
            cell = ws.cell(1, c, h); cell.font = hdr; cell.fill = hdrfill
        for ri, row in enumerate(rows, 2):
            for c, v in enumerate(row, 1):
                cell = ws.cell(ri, c, v); cell.alignment = wrap
            if fills:
                key = next((k for k in _READ_FILL if str(row[fills[1]]).upper().startswith(k)), None)
                if key:
                    ws.cell(ri, fills[0] + 1).fill = PatternFill("solid", fgColor=_READ_FILL[key])
        for c, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(c)].width = w
        ws.freeze_panes = "A2"

    # 1. Matchup Map
    ws1 = wb.active; ws1.title = "Matchup Map"
    sheet(ws1, ["Opponent", "Archetype", "Read", "Posture", "Mission (you)", "Deployment",
                "Biggest threat", "Win condition"],
          [[d["opp"], d["arch"], d["read_short"], d["posture"], d["mission"], d["deployment"] or "",
            (d["priority_kills"][0][0] if d["priority_kills"] else (d["play_around"][0][0] if d["play_around"] else "-")),
            d["wincon"]] for d in P["details"]],
          [30, 14, 20, 10, 20, 14, 22, 52], fills=(2, 2))

    # 2. Threats & Plan
    ws2 = wb.create_sheet("Threats & Plan")
    def lst(items, fmt):
        return "\n".join(fmt(x) for x in items) or "-"
    sheet(ws2, ["Opponent", "Read", "Priority kills (dmg / removable%)", "Play around (survive% / dmg)",
                "Your workhorses", "Your liabilities", "Secondaries: lean", "Secondaries: discard", "The trap"],
          [[d["opp"], d["read_short"],
            lst(d["priority_kills"], lambda x: f"{x[0]} — {x[1]}w, {x[2]}% kill"),
            lst(d["play_around"], lambda x: f"{x[0]} — {x[1]}% lives, {x[2]}w"),
            lst(d["workhorses"], lambda x: f"{x[0]} — {x[1]}w"),
            lst(d["liabilities"], lambda x: f"{x[0]} — {x[1]}%"),
            "\n".join(d["sec_lean"]) or "-", "\n".join(d["sec_avoid"]) or "-", d["trap"]]
           for d in P["details"]],
          [26, 18, 34, 30, 26, 22, 26, 24, 46], fills=(1, 1))
    for ws in (ws2,):
        for r in range(2, ws.max_row + 1):
            ws.row_dimensions[r].height = 92
    wb.save(path)


def build(me_name="custodes", games=250, opt_opp="necrons", seed=11, pdf=True, xlsx=True):
    P = gather(me_name, games, opt_opp, seed)
    d = os.path.join(os.path.dirname(os.path.dirname(os.path.dirname(os.path.dirname(
        os.path.abspath(__file__))))), "reports")
    os.makedirs(d, exist_ok=True)
    md_path = os.path.join(d, f"dossier-{me_name}.md")
    open(md_path, "w").write(render_markdown(P))
    pdf_path = xlsx_path = None
    if pdf:                                            # RUNBOOK -> PDF
        html_path = os.path.join(d, f"dossier-{me_name}.html")
        open(html_path, "w", encoding="utf-8").write(render_html(P))
        try:
            subprocess.run(["soffice", "--headless", "--convert-to", "pdf", "--outdir", d, html_path],
                           check=True, capture_output=True, timeout=180)
            pdf_path = os.path.join(d, f"dossier-{me_name}.pdf")
        except Exception as ex:
            print(f"(pdf render skipped: {ex})")
    if xlsx:                                           # ANALYSIS -> Excel
        try:
            xlsx_path = os.path.join(d, f"analysis-{me_name}.xlsx")
            render_xlsx(P, xlsx_path)
        except Exception as ex:
            print(f"(xlsx render skipped: {ex})"); xlsx_path = None
    return render_markdown(P), md_path, pdf_path, xlsx_path


def main():
    import argparse
    ap = argparse.ArgumentParser(description="Generate the full tournament dossier for a list.")
    ap.add_argument("me", nargs="?", default="custodes")
    ap.add_argument("--games", type=int, default=250)
    ap.add_argument("--opt", default="necrons")
    a = ap.parse_args()
    doc, md_path, pdf_path, xlsx_path = build(a.me, a.games, a.opt)
    print(doc)
    print(f"\n[markdown: {md_path}]" + (f"\n[Runbook PDF: {pdf_path}]" if pdf_path else "")
          + (f"\n[Analysis Excel: {xlsx_path}]" if xlsx_path else ""))


if __name__ == "__main__":
    main()
